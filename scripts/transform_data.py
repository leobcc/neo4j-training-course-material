"""Transform I&D HDP skill mapping into a single denormalized CSV.
Replaces numeric IDs with realistic Italian names while preserving relationships."""

import openpyxl
import csv
import os
import random

random.seed(42)

SRC = r"C:\Users\leob3\AppData\Local\Temp\skill_mapping_copy.xlsx"
DATA_DIR = r"C:\Users\leob3\OneDrive\Desktop\code\neo4j\data"

wb = openpyxl.load_workbook(SRC)
ws = wb["I&D_DE_Skill Dett"]

# --- 1. Extract skill column definitions ---
skill_names = {}
skill_cats = {}
for cell in ws[1]:
    if cell.value is not None:
        skill_cats[cell.column] = str(cell.value).strip()
for cell in ws[2]:
    if cell.value is not None and cell.column not in (1,2,3,4,5,6,7,8,9,10,21,25,41,55,64,76,86):
        skill_names[cell.column] = str(cell.value).strip().replace("\n", " ")

sep_cols = {21, 25, 41, 55, 64}
# Also exclude column headers that are just numbers (separators)
skill_cols = []
for c in range(11, 76):
    if c in sep_cols or c not in skill_names:
        continue
    name = skill_names[c]
    if name.strip().isdigit():
        continue
    skill_cols.append(c)
print(f"Found {len(skill_cols)} skill columns")

skill_headers = [skill_names[c] for c in skill_cols]

# --- 2. Name mappings ---
first_names_m = [
    "Marco", "Luca", "Andrea", "Alessandro", "Matteo", "Francesco", "Giuseppe",
    "Paolo", "Simone", "Fabio", "Stefano", "Roberto", "Davide", "Antonio",
    "Federico", "Claudio", "Gianluca", "Daniele", "Alberto", "Michele",
    "Riccardo", "Emanuele", "Massimo", "Pietro", "Cristiano", "Valerio",
    "Alessio", "Enrico", "Luigi", "Domenico", "Raffaele", "Vincenzo"
]
first_names_f = [
    "Sara", "Martina", "Anna", "Chiara", "Francesca", "Elena", "Valentina",
    "Giulia", "Marta", "Silvia", "Laura", "Caterina", "Alessandra", "Beatrice",
    "Elisa", "Rosa", "Daniela", "Giorgia", "Veronica", "Ilaria",
    "Serena", "Alice", "Erika", "Monica", "Paola", "Bianca"
]
last_names = [
    "Rossi", "Bianchi", "Romano", "Colombo", "Ricci", "Marino", "Greco",
    "Bruno", "Gallo", "Conti", "De Luca", "Costa", "Giordano", "Mancini",
    "Rizzo", "Lombardi", "Moretti", "Barbieri", "Fontana", "Santoro",
    "Mariani", "Rinaldi", "Caruso", "Ferrari", "Bellini", "Fabbri"
]

project_prefixes = ["Progetto", "Iniziativa", "Programma", "Piattaforma", "Sistema"]
project_names = [
    "Apollo", "Sirio", "Vega", "Orione", "Andromeda", "Cassiopea", "Pegaso",
    "Artemis", "Atena", "Demetra", "Era", "Zeus", "Eolo", "Nettuno",
    "Plutone", "Saturno", "Marte", "Venere", "Giove", "Mercurio",
    "Urano", "Atlante", "Prometeo", "Perseo", "Eracle", "Achille",
    "Ulisse", "Ettore", "Giasone", "Teseo", "Edipo", "Paride",
    "Enea", "Dedalo", "Icaro", "Cerbero", "Chimera", "Fenice",
    "Dragone", "Grifone", "Centauro", "Idra", "Leone", "Lupo",
    "Orsa", "Cigno", "Pavone", "Rondine", "Gabbiano", "Falco",
]

sup_names = ["Alberto Conti", "Valentina Marini"]

# Collect unique values
css_data = {}
all_project_ids = set()
all_sup_ids = set()

for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
    cid_raw = row[0]
    if cid_raw is None or (isinstance(cid_raw, float) and str(cid_raw) == "nan"):
        continue
    cid = int(cid_raw)
    if cid not in css_data:
        gender = row[2] if row[2] else "M"
        fnames = first_names_f if gender == "F" else first_names_m
        css_data[cid] = f"{random.choice(fnames)} {random.choice(last_names)}"

    sup_raw = row[1]
    if sup_raw is not None and not (isinstance(sup_raw, float) and str(sup_raw) == "nan"):
        all_sup_ids.add(int(sup_raw))

    pid_raw = row[8]
    if pid_raw is not None and not (isinstance(pid_raw, float) and str(pid_raw) == "nan"):
        pid = int(pid_raw) if isinstance(pid_raw, (int, float)) else str(pid_raw).strip()
        all_project_ids.add(pid)

# Map supervisors
sup_mapping = {}
for i, sid in enumerate(sorted(all_sup_ids)):
    sup_mapping[sid] = sup_names[i % len(sup_names)]

# Map projects
proj_mapping = {}
random.shuffle(project_names)
sorted_pids = sorted([p for p in all_project_ids if isinstance(p, (int, float))])
sorted_pids += sorted([p for p in all_project_ids if not isinstance(p, (int, float))])
for i, pid in enumerate(sorted_pids):
    prefix = random.choice(project_prefixes)
    proj_mapping[pid] = f"{prefix} {project_names[i % len(project_names)]}"

# --- 3. Normalize proficiency ---
def normalize_prof(val):
    if val is None or str(val).startswith("="):
        return ""
    prof = str(val).strip()
    lower = prof.lower()
    if lower in ("novice", "beginner", "experienced", "master"):
        return lower.capitalize()
    if lower in ("basic", "begineer"):
        return "Beginner"
    if lower in ("expert",):
        return "Master"
    return ""  # skip unparseable

# --- 4. Write single CSV ---
header = [
    "employee_id", "name", "gender", "discipline", "department",
    "code", "location", "title",
    "supervisor_id", "supervisor_name",
    "project_id", "project_name", "role",
] + skill_headers

out_path = os.path.join(DATA_DIR, "org_graph.csv")
with open(out_path, "w", newline="", encoding="utf-8") as f:
    w = csv.writer(f)
    w.writerow(header)

    row_count = 0
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        cid_raw = row[0]
        if cid_raw is None or (isinstance(cid_raw, float) and str(cid_raw) == "nan"):
            continue
        cid = int(cid_raw)

        # Employee fields
        emp_name = css_data[cid]
        gender = row[2] if row[2] else ""
        disc = str(row[3]).strip() if row[3] else ""
        dept = str(row[4]).strip() if row[4] else ""
        code = str(row[5]).strip() if row[5] and not (isinstance(row[5], float) and str(row[5]) == "nan") else ""
        loc = row[6] if row[6] else ""
        title = row[7] if row[7] else ""

        # Supervisor fields
        sup_raw = row[1]
        if sup_raw is not None and not (isinstance(sup_raw, float) and str(sup_raw) == "nan"):
            sup_id = str(int(sup_raw))
            sup_name = sup_mapping[int(sup_raw)]
        else:
            sup_id = ""
            sup_name = ""

        # Project fields
        pid_raw = row[8]
        role = row[9] if row[9] else ""
        if pid_raw is not None and not (isinstance(pid_raw, float) and str(pid_raw) == "nan"):
            pid = int(pid_raw) if isinstance(pid_raw, (int, float)) else str(pid_raw).strip()
            proj_name = proj_mapping[pid]
        else:
            pid = ""
            proj_name = ""

        # Skill columns (wide format)
        skill_vals = []
        for col in skill_cols:
            val = row[col - 1] if col - 1 < len(row) else None
            skill_vals.append(normalize_prof(val))

        out_row = [cid, emp_name, gender, disc, dept, code, loc, title,
                   sup_id, sup_name, pid, proj_name, role] + skill_vals
        w.writerow(out_row)
        row_count += 1

print(f"Written: {out_path}")
print(f"Rows: {row_count}")
print(f"Columns: {len(header)}")
print(f"Unique employees: {len(css_data)}")
print(f"Unique supervisors: {len(sup_mapping)}")
print(f"Unique projects: {len(proj_mapping)}")
